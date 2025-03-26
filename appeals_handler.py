from openpyxl import load_workbook, Workbook

FILE_NAME = "Data_from_Appeals.xlsx"

def save_user_data(user_id, username):
    """Сохраняет TG ID и Username в Excel при первом взаимодействии"""
    try:
        wb = load_workbook(FILE_NAME)
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["TG_ID", "Username", "ФИО", "Обращение", "Населенный пункт", "Тема обращения", "Контактный телефон", "Статус обращения", "Долгота (домашняя)", "Широта (домашняя)", "Долгота (иная)", "Широта (иная)"])
    else:
        sheet = wb.active

    # Находим первую пустую строку
    row_to_add = sheet.max_row + 1

    # Записываем данные пользователя
    sheet.append([user_id, username, "Не указано", "Не указано", "Не указано", "Не указано", "Не указано", "Не указано", "Не указано", "Не указано", "Не указано", "Не указано"])
    wb.save(FILE_NAME)

def update_user_data(user_id, full_name=None):
    """Обновляет данные пользователя в Excel по его TG_ID"""
    try:
        wb = load_workbook(FILE_NAME)
    except FileNotFoundError:
        return  # Если файл не найден, выходим

    sheet = wb.active

    # Ищем строку с нужным TG_ID
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
        tg_id_cell = row[0]  # Первая колонка (TG_ID)
        if tg_id_cell.value == user_id:
            # Обновляем ФИО
            row[2].value = full_name or "Не указано"  # ФИО
            wb.save(FILE_NAME)
            return True  # Успешно обновили

    return False  # Если не нашли пользователя по TG_ID

def update_user_location(user_id, location=None):
    """Обновляет населенный пункт пользователя в Excel по его TG_ID"""
    try:
        wb = load_workbook(FILE_NAME)
    except FileNotFoundError:
        return  # Если файл не найден, выходим

    sheet = wb.active

    # Ищем строку с нужным TG_ID
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
        tg_id_cell = row[0]  # Первая колонка (TG_ID)
        if tg_id_cell.value == user_id:
            # Если в "Населенный пункт" указано "Не указано", заменяем на новое значение
            if row[4].value == "Не указано":
                row[4].value = location or "Не указано"  # Населенный пункт
            wb.save(FILE_NAME)
            return True  # Успешно обновили

    return False  # Если не нашли пользователя по TG_ID

def update_user_appeal(user_id, appeal):
    """Обновляет 'Обращение' в Excel: если 'Не указано' — заменяет, иначе создаёт новую запись"""
    try:
        wb = load_workbook(FILE_NAME)
    except FileNotFoundError:
        return  # Если файл не найден, выходим

    sheet = wb.active

    # Ищем строки пользователя по TG_ID
    user_rows = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
        if row[0].value == user_id:  # Сравниваем с TG_ID
            user_rows.append(row)

    if not user_rows:
        return False  # Если не нашли пользователя по TG_ID

    first_row = user_rows[0]  # Первая запись пользователя в файле

    # Если в первой строке "Обращение" == "Не указано", обновляем его
    if first_row[3].value == "Не указано":
        first_row[3].value = appeal
    else:
        # Создаём новую строку с изменённым значением в "Обращение"
        new_row = [cell.value for cell in first_row]
        new_row[3] = appeal  # Меняем только колонку "Обращение"
        sheet.append(new_row)  # Добавляем новую строку

    wb.save(FILE_NAME)
    return True

# Функция для обновления контактного телефона в Excel
def update_contact_phone(user_id, phone_value):
    """Обновляет контактный телефон пользователя в Excel по его TG_ID, но только для строки с 'Запись к специалистам' в 'Обращение'"""
    try:
        wb = load_workbook(FILE_NAME)
    except FileNotFoundError:
        return  # Если файл не найден, выходим

    sheet = wb.active

    # Ищем строку с нужным TG_ID и "Запись к специалистам" в колонке "Обращение"
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
        tg_id_cell = row[0]  # Первая колонка (TG_ID)
        appeal_cell = row[3]  # Четвертая колонка (Обращение)

        if tg_id_cell.value == user_id and appeal_cell.value == "Запись к специалистам":
            # Обновляем колонку "Контактный телефон"
            row[6].value = phone_value  # Колонка "Контактный телефон" (индекс 6)
            wb.save(FILE_NAME)
            return True  # Успешно обновили

    return False  # Если не нашли строку с нужным TG_ID и "Запись к специалистам"