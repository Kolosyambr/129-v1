import telebot
import openpyxl
from telebot.types import Message, ReplyKeyboardMarkup, KeyboardButton
from docx import Document
from math import radians, sin, cos, sqrt, atan2

TOKEN = "7662734466:AAFeIkLV6zkROToipNUU_DdBJV8ge_WRgKM"
bot = telebot.TeleBot(TOKEN)

user_data = {}

FILE_NAME = "Usrs_Data.xlsx"
APPEALS_DATA_FILE = "Data_from_Appeals.xlsx"
PHONE_BOOK_FILE = "Phones.docx"
LINKS_BOOK_FILE = "Ссылки_на_сервисы.docx"

# Флаг для отслеживания этапа регистрации
registration_stage = {}

def check_user_registration(user_id, username):
    """Проверяет наличие пользователя в файле и, если его нет, добавляет в файл."""
    try:
        wb = openpyxl.load_workbook(FILE_NAME)
        sheet = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["TG_ID", "Username", "ФИО", "Населенный пункт", "Контактный телефон", "Долгота (домашняя)", "Широта (домашняя)", "Долгота (иная)", "Широта (иная)"])  # Заголовки
    
    # Проверка наличия пользователя по TG_ID
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:  # Если TG_ID найден
            return True, sheet  # Пользователь найден
    
    # Если TG_ID не найден, добавляем нового пользователя
    sheet.append([user_id, username, "", "", "", "", "", "", ""])  # Записываем пользователя с пустыми данными
    wb.save(FILE_NAME)
    return False, sheet

def save_data_to_excel(user_id, column, data):
    """Сохраняет данные в соответствующую колонку пользователя по TG_ID."""
    try:
        wb = openpyxl.load_workbook(FILE_NAME)
        sheet = wb.active
    except FileNotFoundError:
        return  # Файл не найден, ничего не сохраняем
    
    # Находим строку по TG_ID
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[0].value == user_id:
            row[column].value = data  # Сохраняем данные в нужную колонку
            wb.save(FILE_NAME)
            break

def load_document(file_path):
    """Загружает содержимое .docx файла и возвращает его как строку."""
    doc = Document(file_path)
    content = []
    for para in doc.paragraphs:
        content.append(para.text)
    return "\n".join(content)

def save_user_appeal(user_id, appeal_name, appeal_topic):
    """Записывает данные в файл Data_from_Appeals.xlsx."""
    try:
        # Открываем файл с данными пользователей (Usrs_Data.xlsx)
        wb_user_data = openpyxl.load_workbook(FILE_NAME)
        sheet_user_data = wb_user_data.active
    except FileNotFoundError:
        return  # Если файл не найден, ничего не сохраняем

    # Ищем строку с данным TG_ID в Usrs_Data.xlsx
    user_data = None
    for row in sheet_user_data.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:
            user_data = row
            break
    
    if not user_data:
        return  # Если пользователь не найден, ничего не делаем

    # Открываем файл с обращениями (Data_from_Appeals.xlsx)
    try:
        wb_appeals_data = openpyxl.load_workbook(APPEALS_DATA_FILE)
        sheet_appeals_data = wb_appeals_data.active
    except FileNotFoundError:
        return  # Если файл не найден, ничего не сохраняем

    # Находим первую пустую строку в Data_from_Appeals.xlsx
    new_row = sheet_appeals_data.max_row + 1

    # Колонки, которые нужно заполнить данными из Usrs_Data.xlsx (все, кроме "Обращение", "Тема обращения", "Статус обращения")
    columns_to_copy = ["TG_ID", "Username", "ФИО", "Населенный пункт", "Контактный телефон", "Долгота (домашняя)", "Широта (домашняя)", "Долгота (иная)", "Широта (иная)"]
    column_mapping = {
        "TG_ID": 0,
        "Username": 1,
        "ФИО": 2,
        "Населенный пункт": 3,
        "Контактный телефон": 4,
        "Долгота (домашняя)": 5,
        "Широта (домашняя)": 6,
        "Долгота (иная)": 7,
        "Широта (иная)": 8,
    }

    # Копируем данные из Usrs_Data.xlsx в Data_from_Appeals.xlsx в соответствующие колонки
    for column_name, column_index in column_mapping.items():
        if column_index < len(user_data):  # Проверяем, что данные есть для этой колонки
            sheet_appeals_data.cell(row=new_row, column=column_index + 1, value=user_data[column_index])

    # Заполняем дополнительные данные
    sheet_appeals_data.cell(row=new_row, column=10, value=appeal_name)  # Обращение
    sheet_appeals_data.cell(row=new_row, column=11, value=appeal_topic)  # Тема обращения
    sheet_appeals_data.cell(row=new_row, column=12, value="Обращение закрыто без привлечения оператора")  # Статус обращения

    # Сохраняем изменения в файл Data_from_Appeals.xlsx
    wb_appeals_data.save(APPEALS_DATA_FILE)

@bot.message_handler(func=lambda message: message.text == "Телефонный справочник")
def handle_phone_book(message):
    """Обрабатывает нажатие кнопки "Телефонный справочник"."""
    appeal_name = "Телефонный справочник"
    appeal_topic = "Предоставление телефонного справочника"
    user_id = message.from_user.id

    # Открываем и читаем содержимое документа
    phone_book_content = load_document(PHONE_BOOK_FILE)
    
    # Отправляем содержимое пользователю
    bot.send_message(user_id, phone_book_content)

    # Записываем информацию о обращении в Data_from_Appeals.xlsx
    save_user_appeal(user_id, appeal_name, appeal_topic)

@bot.message_handler(func=lambda message: message.text == "Ссылки на сайты")
def handle_links(message):
    """Обрабатывает нажатие кнопки "Ссылки на сайты"."""
    appeal_name = "Ссылки на сервисы"
    appeal_topic = "Предоставление перечня ссылок"
    user_id = message.from_user.id

    # Открываем и читаем содержимое документа
    links_book_content = load_document(LINKS_BOOK_FILE)
    
    # Отправляем содержимое пользователю
    bot.send_message(user_id, links_book_content)

    # Записываем информацию о обращении в Data_from_Appeals.xlsx
    save_user_appeal(user_id, appeal_name, appeal_topic)

@bot.message_handler(func=lambda message: message.text == "Мероприятия СМФЦ для записи")
def handle_events(message):
    """Обрабатывает нажатие кнопки "Мероприятия СМФЦ для записи"."""
    appeal_name = "Мероприятия СМФЦ для записи"
    appeal_topic = "Предоставление списка мероприятий"
    user_id = message.from_user.id

    events_file = "Мероприятия_СМФЦ_для_записи.docx"

    try:
        with open(events_file, "rb") as file:
            bot.send_document(user_id, file)  # Отправляем сам файл пользователю
    except FileNotFoundError:
        bot.send_message(user_id, "Файл с мероприятиями не найден.")

    # Записываем информацию об обращении в Data_from_Appeals.xlsx
    save_user_appeal(user_id, appeal_name, appeal_topic)

@bot.message_handler(func=lambda message: message.text == "Адрес СМФЦ, геолокация")
def handle_smfcs(message):
    """Обрабатывает нажатие кнопки 'Адрес СМФЦ, геолокация'."""
    user_id = message.from_user.id
    markup = ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
    button1 = KeyboardButton("Ближайший СМФЦ к домашнему адресу")
    button2 = KeyboardButton("Ближайший СМФЦ к моему местоположению")
    markup.add(button1, button2)
    bot.send_message(user_id, "Вы хотите узнать адрес ближайшего СМФЦ к Вашему дому или к Вашему текущему местоположению?", reply_markup=markup)

def calculate_distance(lat1, lon1, lat2, lon2):
    """Вычисляет расстояние между двумя точками по формуле Haversine."""
    # Радиус Земли в километрах
    R = 6371.0
    
    lat1 = radians(lat1)
    lon1 = radians(lon1)
    lat2 = radians(lat2)
    lon2 = radians(lon2)
    
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    
    a = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
    c = 2 * atan2(sqrt(a), sqrt(1 - a))
    
    distance = R * c  # Расстояние в километрах
    return distance

@bot.message_handler(func=lambda message: message.text == "Ближайший СМФЦ к домашнему адресу")
def handle_nearest_smfcs_home(message):
    """Обрабатывает нажатие кнопки 'Ближайший СМФЦ к домашнему адресу'."""
    user_id = message.from_user.id
    
    # Открываем файл Usrs_Data.xlsx, чтобы получить долготу и широту
    try:
        wb_user_data = openpyxl.load_workbook(FILE_NAME)
        sheet_user_data = wb_user_data.active
    except FileNotFoundError:
        bot.send_message(user_id, "Ошибка при чтении файла данных пользователя.")
        return

    # Получаем данные пользователя по TG_ID
    user_lat_home = None
    user_lon_home = None
    for row in sheet_user_data.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:
            user_lat_home = row[6]  # Широта (домашняя)
            user_lon_home = row[5]  # Долгота (домашняя)
            break

    if not user_lat_home or not user_lon_home:
        bot.send_message(user_id, "Не удалось найти данные о вашем домашнем адресе.")
        return

    print(f"Найдено местоположение пользователя: {user_lat_home}, {user_lon_home}")

    # Открываем файл SMFCLocations.xlsx для поиска ближайшего СМФЦ
    try:
        wb_smfcs = openpyxl.load_workbook("SMFCLocations.xlsx")
        sheet_smfcs = wb_smfcs.active
    except FileNotFoundError:
        bot.send_message(user_id, "Ошибка при чтении файла с данными СМФЦ.")
        return

    nearest_distance = float("inf")
    nearest_location = None
    for row in sheet_smfcs.iter_rows(min_row=2, values_only=True):
        smfc_lat = row[2]  # Широта СМФЦ
        smfc_lon = row[1]  # Долгота СМФЦ
        smfc_address = row[0]  # Фактический адрес
        smfc_work_mode = row[3]  # Режим работы (с картинкой)

        # Вычисляем расстояние между пользователем и СМФЦ
        distance = calculate_distance(user_lat_home, user_lon_home, smfc_lat, smfc_lon)
        
        if distance < nearest_distance:
            nearest_distance = distance
            nearest_location = (smfc_lat, smfc_lon, smfc_address, smfc_work_mode)

    if nearest_location:
        smfc_lat, smfc_lon, smfc_address, smfc_work_mode = nearest_location
        print(f"Ближайший СМФЦ найден на расстоянии {nearest_distance} км: {smfc_address}")
        print(f"Координаты ближайшего СМФЦ: Широта: {smfc_lat}, Долгота: {smfc_lon}")

        # Отправляем геолокацию ближайшего СМФЦ
        bot.send_location(user_id, smfc_lat, smfc_lon)
        
        # Отправляем фактический адрес и режим работы
        bot.send_message(user_id, f"Фактический адрес: {smfc_address}")
        
        # Отправляем картинку режима работы (предполагаем, что картинка хранится в ячейке как путь к файлу)
        if smfc_work_mode:
            bot.send_photo(user_id, smfc_work_mode)
        else:
            bot.send_message(user_id, "Режим работы не указан.")
    else:
        bot.send_message(user_id, "Не удалось найти ближайший СМФЦ.")

def send_newMain_menu(user_id):
    """Отправляем новое главное меню после завершения регистрации"""
    markup = ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
    markup.add(KeyboardButton("Запись к специалистам"))
    markup.add(KeyboardButton("Телефонный справочник"))
    markup.add(KeyboardButton("Адрес СМФЦ, геолокация"))
    markup.add(KeyboardButton("Ссылки на сайты"))
    markup.add(KeyboardButton("Тел. ОСЗН Тульской области"))
    markup.add(KeyboardButton("Мероприятия СМФЦ для записи"))
    markup.add(KeyboardButton("Консультация диспетчера"))

    bot.send_message(user_id, "Какой вопрос Вас интересует?", reply_markup=markup)


@bot.message_handler(commands=['start'])
def start(message):
    user_id = message.from_user.id
    username = message.from_user.username
    
    user_exists, sheet = check_user_registration(user_id, username)
    
    if user_exists:
        bot.send_message(user_id, "С возвращением, давно не виделись!")
        check_authorization_complete(user_id)
    else:
        bot.send_message(user_id, "Для дальнейшей корректной работы с сервисом пройдите простую регистрацию.")
        # Кнопки для регистрации
        markup = ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
        Regbutton1 = KeyboardButton("ФИО")
        Regbutton2 = KeyboardButton("Населенный пункт")
        Regbutton3 = KeyboardButton("Контактный телефон")
        Regbutton4 = KeyboardButton("Адрес проживания")
        markup.add(Regbutton1, Regbutton2, Regbutton3, Regbutton4)
        bot.send_message(user_id, "Выберите, что хотите заполнить для регистрации:", reply_markup=markup)

@bot.message_handler(func=lambda message: message.text in ["ФИО", "Населенный пункт", "Контактный телефон", "Адрес проживания"])
def handle_registration(message):
    user_id = message.from_user.id
    user_stage = registration_stage.get(user_id)

    # Сброс предыдущего этапа, если была другая кнопка
    if user_stage and user_stage != message.text:
        bot.send_message(user_id, "Вы начали новый этап регистрации. Пожалуйста, предоставьте информацию для следующего поля.")
    
    # Устанавливаем текущий этап регистрации
    registration_stage[user_id] = message.text

    if message.text == "ФИО":
        bot.send_message(user_id, "Пожалуйста, введите Ваше ФИО (В формате Фамилия, Имя, Отчество).")
    elif message.text == "Населенный пункт":
        bot.send_message(user_id, "Пожалуйста, введите Ваш город проживания (Напишите без указания типа населенного пункта).")
    elif message.text == "Контактный телефон":
        bot.send_message(user_id, "Укажите пожалуйста, Ваш актуальный номер телефона. Он будет использоваться для диспетчера с Вами, в случае необходимости.")
    elif message.text == "Адрес проживания":
        markup = ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
        button1 = KeyboardButton("Да")
        button2 = KeyboardButton("Нет")
        markup.add(button1, button2)
        bot.send_message(user_id, "Вы сейчас находитесь дома?", reply_markup=markup)

@bot.message_handler(func=lambda message: message.text == "Да" or message.text == "Нет")
def handle_address(message):
    user_id = message.from_user.id

    if message.text == "Да":
        # Просим отправить геолокацию
        bot.send_message(user_id, "Пожалуйста, отправьте свою текущую геолокацию.")
        bot.register_next_step_handler(message, process_location_home)
    elif message.text == "Нет":
        # Просим выбрать геоточку на карте
        bot.send_message(user_id, "Пожалуйста, выберите геоточку на карте.")
        bot.register_next_step_handler(message, process_location_other)

@bot.message_handler(content_types=['location'])
def process_location_home(message):
    """Обрабатывает отправку геолокации пользователем (для "Да", когда он дома)."""
    user_id = message.from_user.id
    if message.location:
        # Сохраняем долготу и широту в соответствующие колонки
        save_data_to_excel(user_id, 5, message.location.longitude)  # Долгота (домашняя)
        save_data_to_excel(user_id, 6, message.location.latitude)   # Широта (домашняя)
        bot.send_message(user_id, "Ваша геолокация успешно сохранена.")
        check_registration_complete(user_id)

def process_location_other(message):
    """Обрабатывает выбор геоточки на карте (для "Нет", когда не дома)."""
    user_id = message.from_user.id
    if message.location:
        # Сохраняем долготу и широту в соответствующие колонки
        save_data_to_excel(user_id, 7, message.location.longitude)  # Долгота (иная)
        save_data_to_excel(user_id, 8, message.location.latitude)   # Широта (иная)
        bot.send_message(user_id, "Ваша геоточка успешно сохранена.")
        check_registration_complete(user_id)

@bot.message_handler(func=lambda message: message.text not in ["ФИО", "Населенный пункт", "Контактный телефон", "Адрес проживания"])
def handle_text_input(message):
    user_id = message.from_user.id
    user_stage = registration_stage.get(user_id)

    # Если этап не найден, игнорируем сообщение
    if not user_stage:
        return
    
    # Сохраняем данные в Excel в зависимости от текущего этапа
    if user_stage == "ФИО":
        save_data_to_excel(user_id, 2, message.text)  # Сохраняем в колонку "ФИО"
    elif user_stage == "Населенный пункт":
        save_data_to_excel(user_id, 3, message.text)  # Сохраняем в колонку "Населенный пункт"
    elif user_stage == "Контактный телефон":
        save_data_to_excel(user_id, 4, message.text)  # Сохраняем в колонку "Контактный телефон"
    
    # После записи данных, проверяем, все ли данные заполнены
    check_registration_complete(user_id)

def check_registration_complete(user_id):
    """Проверяет, все ли данные заполнены для текущего пользователя"""
    try:
        wb = openpyxl.load_workbook(FILE_NAME)
        sheet = wb.active
    except FileNotFoundError:
        return  # Файл не найден, ничего не проверяем

    # Проверяем заполненность данных
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:
            if all([cell is not None for cell in row[2:7]]):  # Проверяем, что данные в колонках 3-7 заполнены
                bot.send_message(user_id, "Спасибо за предоставленную информацию, данные успешно записаны!")
                # Отправляем новое меню
                send_newMain_menu(user_id)
                break

def check_authorization_complete(user_id):
    """Проверяет, все ли данные заполнены для текущего пользователя"""
    try:
        wb = openpyxl.load_workbook(FILE_NAME)
        sheet = wb.active
    except FileNotFoundError:
        return  # Файл не найден, ничего не проверяем

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:
            if all([cell is not None for cell in row[2:7]]):
                send_newMain_menu(user_id)
                break

def send_newMain_menu(user_id):
    """Отправляем новое меню после завершения регистрации"""
    markup = ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
    markup.add(KeyboardButton("Запись к специалистам"))
    markup.add(KeyboardButton("Телефонный справочник"))
    markup.add(KeyboardButton("Адрес СМФЦ, геолокация"))
    markup.add(KeyboardButton("Ссылки на сайты"))
    markup.add(KeyboardButton("Тел. ОСЗН Тульской области"))
    markup.add(KeyboardButton("Мероприятия СМФЦ для записи"))
    markup.add(KeyboardButton("Консультация диспетчера"))

    bot.send_message(user_id, "Какой вопрос Вас интересует?", reply_markup=markup)

if __name__ == "__main__":
    bot.infinity_polling()
